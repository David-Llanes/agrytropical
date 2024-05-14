import os
from dotenv import load_dotenv
from supabase import create_client, Client
import random
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook

_instance = None

def get_supabase_instance():
    global _instance
    if _instance is None:
        _instance = SupabaseDataLayer()
    return _instance

def desenpaquetar_set(data):
    setxd = set(data)
    return setxd.pop()

def calcular_total_usd(dato):
    precio = dato['VariedadMango']['precio']
    cajas = dato['cajas']
    return round(precio * cajas, 2)

def calcular_total_mxn(dato, tipo_cambio):
    precio = dato['VariedadMango']['precio']
    cajas = dato['cajas']
    return round(precio * cajas * tipo_cambio, 2)


class SupabaseDataLayer:
    def __init__(self):
        load_dotenv()
        # Configura las credenciales de tu aplicación Supabase
        supabase_url = os.environ.get("SUPABASE_URL")
        supabase_key = os.environ.get("SUPABASE_KEY")
        
        # Crea un cliente Supabase
        self.client: Client = create_client(supabase_url, supabase_key)

    def authenticate(self, email, password):
        # Método para autenticar un usuario
        user = self.client.auth.sign_in_with_password({ "email": email, "password": password })
        if user:
            return user
        else:
            raise Exception(f"Error en la autenticación.")
        
    def get_resumen(self, id):
        # Método para obtener
        data = self.client.table("ResumenCarga").select("*").eq("no_certificate", id).execute()
        
        # Assert we pulled real data.
        if len(data.data) < 1:
            return []
        
        print (data.data)
        return data.data
        
    def get_variedades(self):
        data = self.client.table("VariedadMango").select("variedad").execute()
        
        # Assert we pulled real data.
        if len(data.data) < 1:
            return []
        
        array = [element['variedad'] for element in data.data]
        
        # print(array)
        return array
    
    def get_info_manifiesto(self, id):
        infoManifiesto = self.client.table("ManifiestoCarga").select("*").eq("no_certificate", id).execute()
        infoCertificado = self.client.table("CertificadoDeEmbarque").select("*").eq("no_certificate", id).execute()


        if len(infoManifiesto.data) < 1:
            return {}
        
        if len(infoCertificado.data) < 1:
            return {}
        
        infoManifiesto = infoManifiesto.data[0] | infoCertificado.data[0]
        
        print(infoManifiesto)
        return infoManifiesto

    def get_manifiestos(self):
        data = self.client.table("ManifiestoCarga").select("no_certificate").execute()
        
        # Assert we pulled real data.
        if len(data.data) < 1:
            return []
        
        array = [element['no_certificate'] for element in data.data]
        
        # print(array)
        return array

    def insert_manifiesto(self, data, embarque, temp=0):
        noCertificate = data['no_certificate']
        
        try:
            # Inserta un manifiesto
            result = self.client.table("CertificadoDeEmbarque").insert(data).execute()
            
            if result.data and len(result.data) > 0:
                result2 = self.client.table("ManifiestoCarga").insert({ "no_certificate": noCertificate, "no_embarque": embarque, "temp": temp }).execute()
            if result2.data and len(result2.data) < 1:
                raise Exception ("Error al insertar el manifiesto.")
            
            allInfo = self.get_info_manifiesto(noCertificate)
                  
        except Exception as e:
            print(f"Error: {e}")
            return []
        

        return allInfo
    
    def modify_manifiesto(self, data, embarque, temp=0):
        noCertificate = data['no_certificate']
        
        try:
            # Modifica un manifiesto
            result = self.client.table("CertificadoDeEmbarque").update(data).eq("no_certificate", noCertificate).execute()
            
            if result.data and len(result.data) > 0:
                result2 = self.client.table("ManifiestoCarga").update({ "no_certificate": noCertificate, "no_embarque": embarque, "temp": temp }).eq("no_certificate", noCertificate).execute()
            if result2.data and len(result2.data) < 1:
                raise Exception ("Error al modificar el manifiesto.")
        except Exception as e:
            print(f"Error: {e}")
            return []
        
        allInfo = self.get_info_manifiesto(noCertificate)
        
        return allInfo
    
    def insert_resumen(self, data):
        try:
            # Inserta un resumen de carga
            result = self.client.table("ResumenCarga").insert(data).execute()
            
            if result.data and len(result.data) < 1:
                raise Exception ("Error al insertar el resumen.")
            
            info = self.get_resumen(data['no_certificate'])
            return info
        except Exception as e:
            print(f"Error: {e}")
            return []
    
    def obtener_tipoDeCambio(self, fechaCarga):
        fecha = datetime.strptime(fechaCarga, '%Y-%m-%d')
        seed = int(fecha.strftime('%Y%m%d'))
        random.seed(seed)
        
        tipoCambio = round(random.uniform(16.50, 17.20), 2)
        print (f"Tipo de cambio: {tipoCambio}")
        return tipoCambio
    
    def get_costos(self, id):
        # Método para obtener
        data = self.client.table("ResumenCarga").select("variedad, calibre,cajas, VariedadMango(precio), CertificadoDeEmbarque(fecha_carga)").eq("no_certificate", id).execute()
        
        # Assert we pulled real data.
        if len(data.data) < 1:
            return []
        
        print(data.data)
        
        fecha = data.data[0]['CertificadoDeEmbarque']['fecha_carga']
        tipoCambio = self.obtener_tipoDeCambio(fecha)
        
        # Sacar la cantidad de cajas por variedad
        cajasPorVariedad = { dato['variedad']: {'cajas': sum([elem['cajas'] for elem in data.data if elem['variedad'] == dato['variedad']]), 'precio': desenpaquetar_set([dato['VariedadMango']['precio'] for elem in data.data if elem['variedad'] == dato['variedad']])} for dato in data.data }
        
        # Sacar costos por variedad
        costosPorVariedad = {variedad: {'cajas': cajasPorVariedad[variedad]['cajas'], 'precio': cajasPorVariedad[variedad]['precio'], 'total_usd': round(cajasPorVariedad[variedad]['cajas'] * cajasPorVariedad[variedad]['precio'],2), 'total_mxn': round(cajasPorVariedad[variedad]['cajas'] * cajasPorVariedad[variedad]['precio'] * tipoCambio, 2)} for variedad in cajasPorVariedad}
        
        
        for dato in data.data:
            del(dato['CertificadoDeEmbarque'])
            dato['total_usd'] = calcular_total_usd(dato)
            dato['total_mxn'] = calcular_total_mxn(dato, tipoCambio)
        
        resultado = {
            "no_certificate": id,
            "fecha_carga": fecha,
            "tipo_cambio": tipoCambio,
            "costos": data.data,
            "costo_total_usd": sum([dato['total_usd'] for dato in data.data]),
            "costo_total_mxn": sum([dato['total_mxn'] for dato in data.data]),
            "costosPorVariedad": costosPorVariedad,
        }
        
        # print(resultado)
        return resultado
    
    
    def download_costos(self, id: int):
        data = self.get_costos(id)
        
        if not data:
            return None
        
        noCertificate = data['no_certificate']
        fechaCarga = data['fecha_carga']

        costos = data['costos']
        for dato in costos:
            dato['Variedad'] = dato['variedad']
            dato['Calibre'] = dato['calibre']
            dato['Cajas'] = dato['cajas']
            dato['Precio'] = dato['VariedadMango']['precio']
            dato['Total USD'] = dato['total_usd']
            dato['Total MXN'] = dato['total_mxn']
            
            del(dato['VariedadMango'])
            del(dato['variedad'])
            del(dato['calibre'])
            del(dato['cajas'])
            del(dato['total_usd'])
            del(dato['total_mxn'])
        
        costosPorVariedad = data['costosPorVariedad']

        # Crear un DataFrame para la sección de costos
        costos_df = pd.DataFrame(costos)

        # Crear un DataFrame para costos por variedad
        costos_variedad_df = pd.DataFrame(costosPorVariedad).transpose().reset_index().rename(columns={'index': 'Variedad', 'cajas': 'Cajas', 'precio': 'Precio', 'total_usd': 'Total USD', 'total_mxn': 'Total MXN'})

        # Crear un DataFrame para los valores generales
        general_df = pd.DataFrame({
            'No. Certificado': [data['no_certificate']],
            'Fecha de Carga': [data['fecha_carga']],
            'Tipo de cambio': [data['tipo_cambio']],
            'Costo total en USD': [data['costo_total_usd']],
            'Costo total en MXN': [data['costo_total_mxn']]
        })

        direccion = f"D:/Agrytropical/CostosPorMaquila/{fechaCarga}_{noCertificate}_costos.xlsx"
        
        # # Crear un archivo Excel con los DataFrames en diferentes hojas
        # with pd.ExcelWriter(f"{fechaCarga}_{noCertificate}_costos.xlsx", engine='openpyxl') as writer:
        #     general_df.to_excel(writer, index=False, sheet_name='General')
        #     costos_df.to_excel(writer, index=False, sheet_name='Costos')
        #     costos_variedad_df.to_excel(writer, index=False, sheet_name='Costos por Variedad')
        
        # Crear un archivo Excel con las tres tablas en la misma hoja
        with pd.ExcelWriter(direccion, engine='openpyxl') as writer:
            # Escribir valores generales
            general_df.to_excel(writer, index=False, sheet_name='Costos por maquila', startrow=2)
            
            # Escribir tabla de costos
            startrow = len(general_df) + 4  # Dejar 3 filas de espacio
            inicioTabla2 = f"A{startrow + 1}"
            costos_df.to_excel(writer, index=False, sheet_name='Costos por maquila', startrow=startrow + 1)
            
            # Escribir tabla de costos por variedad
            startrow = startrow + len(costos_df) + 4  # Dejar 3 filas de espacio
            inicioTabla3 = f"A{startrow + 1}"
            costos_variedad_df.to_excel(writer, index=False, sheet_name='Costos por maquila', startrow=startrow + 1)

        # Cargar el archivo Excel para agregar títulos
        wb = load_workbook(direccion)
        ws = wb['Costos por maquila']

        # Agregar títulos encima de cada tabla
        ws['A1'] = f"Costos por maquila: {noCertificate}"
        ws['A2'] = 'Resumen de costos:'
        ws[f"{inicioTabla2}"] = 'Costos por variedad y calibre:'
        ws[f"{inicioTabla3}"] = 'Costos por variedad:'

        # Guardar el archivo Excel
        wb.save(direccion)

        # print(f"Archivo Excel guardado en 'D:/Agrytropical/CostosPorMaquila/{fechaCarga}_{noCertificate}_costos.xlsx")
        return f"Archivo Excel guardado en 'D:/Agrytropical/CostosPorMaquila/{fechaCarga}_{noCertificate}_costos.xlsx"
    
    # Método para descargar el resumen de carga
    def download_resumen(self, id: int):
        manifiesto = self.get_info_manifiesto(id)
        
        if not manifiesto:
            return None
        
        noCertificate = manifiesto['no_certificate']
        fechaCarga = manifiesto['fecha_carga']

        # Crear un DataFrame para el manifiesto
        infoManifiesto = pd.DataFrame({
            'No. Certificado': [manifiesto['no_certificate']],
            'No. Factura': [manifiesto['no_factura']],
            'No. Embarque': [manifiesto['no_embarque']],
            'Fecha de Carga': [manifiesto['fecha_carga']],
            'Temperatura': [manifiesto['temp']],
            'Puerto de salida': [manifiesto['export_port']],
            'Puerto de llegada': [manifiesto['entry_port']],
            'Empaque': [manifiesto['empaque']],
            'Total de pallets': [manifiesto['total_pallets']],
        })
        
        # Crear un DataFrame para el resumen de la carga
        resumen = self.get_resumen(id)
        resumen_df = pd.DataFrame(resumen).rename(columns={'id': 'ID', 'no_certificate': 'No. Certificado', 'marca_caja': 'Marca de caja', 'calibre': 'Calibre', 'variedad': 'Variedad', 'cajas': 'Cajas', 'calidad': 'Calidad'})

        direccion = f"D:/Agrytropical/InformacionCarga/{fechaCarga}_{noCertificate}_resumenCarga.xlsx"
        
        # Crear un archivo Excel con las tres tablas en la misma hoja
        with pd.ExcelWriter(direccion, engine='openpyxl') as writer:
            # Escribir valores generales
            infoManifiesto.to_excel(writer, index=False, sheet_name='Resumen de carga', startrow=2)
            
            # Escribir tabla de costos
            startrow = len(infoManifiesto) + 4  # Dejar 3 filas de espacio
            inicioTabla2 = f"A{startrow + 1}"
            resumen_df.to_excel(writer, index=False, sheet_name='Resumen de carga', startrow=startrow + 1)

        # Cargar el archivo Excel para agregar títulos
        wb = load_workbook(direccion)
        ws = wb['Resumen de carga']

        # Agregar títulos encima de cada tabla
        ws['A1'] = f"Resumen de carga: {noCertificate}"
        ws['A2'] = 'Información relacionada'
        ws[f"{inicioTabla2}"] = 'Cantidad de cajas por variedad y calibre'

        # Guardar el archivo Excel
        wb.save(direccion)

        # print(f"Archivo Excel guardado en {direccion}")
        return f"Archivo Excel guardado en {direccion}"
    
        
        
if __name__ == "__main__":
    test = SupabaseDataLayer()
    # test.download_resumen(4379065)



